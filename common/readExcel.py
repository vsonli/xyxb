# -*- coding:utf-8 -*-
# @Time   :2019/6/4 18:52
# @File   :readExcel.py
# @Author :Vsonli
import openpyxl
from regis import register as res

class Case:
    def __init__(self, attrs):
        """
        初始化用例
        :param attrs: zip类型-->[(key,value),(key1,value1)....]
        """
        for item in attrs:
            setattr(self, item[0], item[1])

class Read_r_excel(object):
    def __init__(self,file_name,sheet_name):
        self.wb = openpyxl.load_workbook(file_name)
        self.sheet = self.wb[sheet_name]
        self.file_name = file_name

    def __del__(self):
        self.wb.close()

    def read_data_line(self):
        rows_data = list(self.sheet.rows)
        titles = []
        for title in rows_data[0]:
            titles.append(title.value)

        cases = []
        #遍历每一条用例
        for case in rows_data[1:]:
            # 获得每一条用例里面的数据
            data = []
            for cell in case:
                if isinstance(cell.value,str):
                    data.append(eval(cell.value))
                else:
                    data.append(cell.value)

            #获得数据后进行打包并放入大列表中
            case_data = dict(list(zip(titles,data)))
            cases.append(case_data)
        return cases

    def read_data_obj(self):
        rows_data = list(self.sheet.rows)
        titles = []
        for title in rows_data[0]:
            titles.append(title.value)

        cases = []
        for case in rows_data[1:]:
            #创建一个Case类的对象存储用例数据
            case_obj = Case()
            data = []
            for cell in case:
                if isinstance(cell.value, str):
                    data.append(eval(cell.value))
                else:
                    data.append(cell.value)

            # 获得数据后进行打包并放入大列表中
            case_data = list(zip(titles, data))
            for item in case_data:
                setattr(case_obj,item[0],item[1])
            # print(case_obj.case_id,case_obj.data,case_obj.excepted)
            cases.append(case_obj)
        return cases

    def r_data(self):
        """
                按行读取excel中的数据，以列表的形式返回，列表中每个对象为一条用例
                excel中的表头为对象的属性，对应的数据为属性值
                :return: type:list--->[case_obj1,case_obj2....]，
                """
        # 按行获取数据转换成列表
        rows_data = list(self.sheet.rows)
        # 获取表单的表头信息
        titles = []
        for title in rows_data[0]:
            titles.append(title.value)
        # 定义一个空列表用来存储所有的用例
        cases = []
        # 判断是否是读取所有数据
        for case in rows_data[1:]:
            # data用例临时存放用例数据
            data = []
            # 判断该单元格是否为字符串类型，
            for cell in case:
                data.append(cell.value)
            # 将该条数据放入cases中
            case_data = zip(titles, data)
            # 创建一个Case类的对象，用来保存用例数据，
            case_obj = Case(case_data)
            cases.append(case_obj)
        return cases

    def r_data_obj(self, list1):
        '''做了下改变,因为从配置文件读过来的是字符串，需要eval转换下，然后判断是否为空'''
        # list1 = eval(list1)
        # if list1 == None:
        #     return self.r_data()

        # 获取最大行
        max_r = self.sheet.max_row
        # 定义一个空列表，存储所有的用例
        cases = []
        titles = []
        for row in range(1, max_r + 1):
            if row == 1:  # 判断是否第一行，如果是，则获取表头
                for column in list1:  # 遍历所有列
                    # 获取表格内的内容
                    title = self.sheet.cell(row, column).value
                    titles.append(title)
            else:  # 获取剩余的用例数据
                case_info = []  # 定义空列表存储剩余每一行的数据
                for column in list1:
                    info = self.sheet.cell(row, column).value
                    # 把表格中的数据，添加到列表中
                    case_info.append(info)

                # 把表头和数据打包，放入字典
                case = list(zip(titles, case_info))
                # 将一条用例存入一个对象中（每一列对应对象的一个属性）
                case_obj = Case(case)
                for i in case:
                    setattr(case_obj, i[0], i[1])
                # 把字典加入cases中
                cases.append(case_obj)
        return cases

    def write_data(self, row, column, msg):
        '''写入数据'''
        print('我被调用了')
        self.sheet.cell(row=row, column=column, value=msg)
        self.wb.save(self.file_name)


if __name__ == '__main__':
    from common.apicom import *
    tokens = get_token()
    head = get_head(tokens)
    r = Read_r_excel('D://Vson//xyxb//data//test1.xlsx','Sheet')
    vals = r.r_data_obj([1,2,3])
    couse = get_course('',head,'',1)
    for item in vals:
        result = res(item.url,head,item.data,1)
        print(result)